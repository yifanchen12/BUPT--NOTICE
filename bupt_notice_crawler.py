# -*- coding: utf-8 -*-
"""
北京邮电大学信息门户“校内通知”活动整理爬虫。

登录策略：
1. 首次运行会打开系统 Chrome。
2. 如果出现 CAS 登录页，请在浏览器里正常登录。
3. 程序会把登录态保存在 runtime/chrome-profile，后续运行会自动复用。
"""

from __future__ import annotations

import argparse
import datetime as dt
import os
import re
import sys
import time
import urllib.error
import urllib.request
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, List, Optional, Sequence, Tuple
from urllib.parse import urljoin

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
from playwright.sync_api import sync_playwright


DEFAULT_HOME_URL = "http://my.bupt.edu.cn/xs_index.jsp?urltype=tree.TreeTempUrl&wbtreeid=1541"
DEFAULT_LIST_URL = "http://my.bupt.edu.cn/list.jsp?urltype=tree.TreeTempUrl&wbtreeid=1154"

EVENT_KEYWORDS = (
    "活动",
    "比赛",
    "竞赛",
    "讲座",
    "报告",
    "培训",
    "会议",
    "论坛",
    "沙龙",
    "宣讲",
    "报名",
    "征集",
    "体验",
    "课程",
    "团体赛",
    "志愿",
    "实践",
    "节",
)

CAMPUS_LOCATION_KEYWORDS = (
    "北京邮电大学",
    "北邮",
    "校内",
    "校区",
    "西土城",
    "沙河",
    "宏福",
    "小西天",
    "教一",
    "教二",
    "教三",
    "教四",
    "教学楼",
    "主楼",
    "科研楼",
    "经管楼",
    "行政楼",
    "学生发展中心",
    "学生活动中心",
    "科学会堂",
    "体育馆",
    "图书馆",
    "报告厅",
    "会议室",
    "教室",
    "礼堂",
    "操场",
    "食堂",
)

NON_CAMPUS_LOCATION_KEYWORDS = (
    "线上",
    "在线",
    "腾讯会议",
    "会议号",
    "zoom",
    "直播",
    "校外",
    "清华大学",
    "北京大学",
    "中国人民大学",
    "北京师范大学",
    "北京交通大学",
    "北京科技大学",
    "中央财经大学",
    "中国科学院",
    "国家会议中心",
    "会展中心",
    "酒店",
    "宾馆",
)

IGNORE_DEPARTMENT_CANDIDATES = {
    "校园通知",
    "正文",
    "最新通知",
    "更多",
    ">>",
    ">",
    "搜索",
}


@dataclass
class NoticeSummary:
    title: str
    url: str
    department: str = ""
    publish_date: str = ""


@dataclass
class ArticleData:
    title: str
    body_text: str
    lines: List[str]
    publish_department: str = ""
    publish_date: str = ""


@dataclass
class ActivityRow:
    location: str
    event_time: str
    department: str
    activity_name: str
    publish_date: str
    notice_title: str
    url: str
    note: str
    sort_time: str


def app_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def default_output_path() -> Path:
    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    return app_dir() / "outputs" / f"校内通知活动_{stamp}.xlsx"


def today_text() -> str:
    return dt.date.today().strftime("%Y-%m-%d")


def parse_cli_date(value: str) -> str:
    value = normalize_date_text(value)
    try:
        return dt.datetime.strptime(value, "%Y-%m-%d").strftime("%Y-%m-%d")
    except ValueError as exc:
        raise argparse.ArgumentTypeError("日期格式应为 YYYY-MM-DD") from exc


def resolve_target_date(args) -> str:
    if args.date:
        return parse_cli_date(args.date)
    if args.today_only:
        return today_text()
    return ""


def normalize_space(text: str) -> str:
    text = text.replace("\u3000", " ").replace("\xa0", " ")
    return re.sub(r"\s+", " ", text).strip()


def clean_line(text: str) -> str:
    text = normalize_space(text)
    return text.strip(" \t\r\n·•▪■●◆◇-—")


def get_text_lines(node) -> List[str]:
    raw = node.get_text("\n", strip=True)
    lines = [clean_line(line) for line in raw.splitlines()]
    return [line for line in lines if line]


def first_date(text: str) -> str:
    patterns = [
        r"20\d{2}[-/.]\d{1,2}[-/.]\d{1,2}",
        r"20\d{2}年\d{1,2}月\d{1,2}日?",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            return normalize_date_text(match.group(0))
    return ""


def normalize_date_text(value: str) -> str:
    match = re.search(r"(20\d{2})[-/.年](\d{1,2})[-/.月](\d{1,2})", value)
    if not match:
        return normalize_space(value)
    year, month, day = map(int, match.groups())
    return f"{year:04d}-{month:02d}-{day:02d}"


def find_chrome_executable(explicit_path: str = "") -> Optional[str]:
    if explicit_path:
        path = Path(explicit_path).expanduser()
        if path.exists():
            return str(path)
        raise FileNotFoundError(f"找不到指定的 Chrome 路径：{path}")

    candidates = [
        Path(os.environ.get("PROGRAMFILES", r"C:\Program Files"))
        / "Google"
        / "Chrome"
        / "Application"
        / "chrome.exe",
        Path(os.environ.get("PROGRAMFILES(X86)", r"C:\Program Files (x86)"))
        / "Google"
        / "Chrome"
        / "Application"
        / "chrome.exe",
        Path(os.environ.get("LOCALAPPDATA", ""))
        / "Google"
        / "Chrome"
        / "Application"
        / "chrome.exe",
        Path(os.environ.get("PROGRAMFILES", r"C:\Program Files"))
        / "Microsoft"
        / "Edge"
        / "Application"
        / "msedge.exe",
        Path(os.environ.get("PROGRAMFILES(X86)", r"C:\Program Files (x86)"))
        / "Microsoft"
        / "Edge"
        / "Application"
        / "msedge.exe",
    ]
    for path in candidates:
        if path.exists():
            return str(path)
    return None


def log(message: str) -> None:
    print(f"[{dt.datetime.now().strftime('%H:%M:%S')}] {message}", flush=True)


def safe_goto(page, url: str, timeout_ms: int = 60000) -> None:
    for attempt in range(1, 4):
        try:
            page.goto(url, wait_until="domcontentloaded", timeout=timeout_ms)
            try:
                page.wait_for_load_state("networkidle", timeout=15000)
            except PlaywrightTimeoutError:
                pass
            return
        except PlaywrightTimeoutError:
            if attempt == 3:
                raise
            log(f"页面加载超时，正在重试第 {attempt + 1} 次：{url}")
            time.sleep(2)


def is_login_page(page) -> bool:
    url = page.url.lower()
    if "authserver" in url or "cas" in url:
        return True
    try:
        title = page.title().lower()
    except Exception:
        title = ""
    if "cas login" in title:
        return True
    try:
        html = page.content()
    except Exception:
        return False
    return "id=\"loginform\"" in html.lower() or "name=\"username\"" in html.lower()


def page_looks_logged_in(page) -> bool:
    try:
        if is_login_page(page):
            return False
        url = page.url.lower()
        if "my.bupt.edu.cn" not in url:
            return False
        html = page.content()
    except Exception:
        return False

    markers = ("退出", "校内通知", "服务门户", "门户首页", "发布部门", "校园通知")
    return any(marker in html for marker in markers)


def cookie_header_for(context, url: str) -> str:
    cookies = context.cookies([url])
    return "; ".join(f"{cookie['name']}={cookie['value']}" for cookie in cookies)


def target_accessible_with_browser_cookies(context, target_url: str, timeout: int = 8) -> bool:
    cookie_header = cookie_header_for(context, target_url)
    if not cookie_header:
        return False

    request = urllib.request.Request(
        target_url,
        headers={
            "Cookie": cookie_header,
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120 Safari/537.36"
            ),
        },
    )
    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            final_url = response.geturl().lower()
            content_type = response.headers.get("Content-Type", "")
            raw = response.read(200000)
    except (urllib.error.URLError, TimeoutError, OSError):
        return False

    if "authserver" in final_url or "cas" in final_url:
        return False
    if "text/html" not in content_type.lower() and not raw:
        return False
    text = raw.decode("utf-8", errors="ignore")
    return "校内通知" in text or "发布部门" in text or "校园通知" in text


def ensure_logged_in(page, target_url: str, login_timeout: int):
    safe_goto(page, target_url)
    if not is_login_page(page):
        return page

    log("检测到登录页。请在打开的浏览器中完成信息门户/CAS 登录，程序会自动继续。")
    deadline = time.time() + login_timeout
    last_probe = 0.0
    last_wait_log = 0.0

    while time.time() < deadline:
        for candidate in page.context.pages:
            if page_looks_logged_in(candidate):
                page = candidate
                log("已检测到登录成功，正在进入校内通知列表。")
                safe_goto(page, target_url)
                return page

        now = time.time()
        if now - last_probe >= 5:
            last_probe = now
            if target_accessible_with_browser_cookies(page.context, target_url):
                log("已检测到浏览器登录态，正在进入校内通知列表。")
                safe_goto(page, target_url)
                if not is_login_page(page):
                    return page

        if now - last_wait_log >= 20:
            last_wait_log = now
            remaining = max(0, int(deadline - now))
            log(f"仍在等待登录完成，剩余 {remaining} 秒。")

        if not is_login_page(page) and "my.bupt.edu.cn" in page.url.lower():
            log("已检测到登录成功。")
            return page

        time.sleep(1)

    raise TimeoutError(
        f"等待登录超时（{login_timeout} 秒）。可用 --login-timeout 调大等待时间。"
    )


def pick_row_container(anchor):
    node = anchor
    fallback = anchor.parent
    for _ in range(8):
        if node is None:
            break
        text = normalize_space(node.get_text(" ", strip=True))
        if node.name in {"li", "tr"}:
            return node
        if first_date(text) and len(text) < 500:
            return node
        node = node.parent
    return fallback or anchor


def infer_department(container, title: str, publish_date: str) -> str:
    chunks = [clean_line(part) for part in container.stripped_strings]
    candidates: List[str] = []
    title_short = title[:18]
    for chunk in chunks:
        if not chunk or chunk in IGNORE_DEPARTMENT_CANDIDATES:
            continue
        if chunk == title or title_short in chunk or chunk in title:
            continue
        if publish_date and publish_date in normalize_date_text(chunk):
            continue
        if first_date(chunk):
            continue
        if len(chunk) > 45:
            continue
        if re.fullmatch(r"[\d\s/.-]+", chunk):
            continue
        candidates.append(chunk)
    return candidates[-1] if candidates else ""


def extract_notice_summaries(html: str, base_url: str) -> List[NoticeSummary]:
    soup = BeautifulSoup(html, "html.parser")
    for tag in soup(["script", "style", "noscript"]):
        tag.decompose()

    results: List[NoticeSummary] = []
    seen_urls = set()
    for anchor in soup.find_all("a", href=True):
        href = anchor.get("href", "").strip()
        title = normalize_space(anchor.get_text(" ", strip=True))
        if not href or len(title) < 4:
            continue

        full_url = urljoin(base_url, href)
        lower = full_url.lower()
        is_content = (
            "xntz_content.jsp" in lower
            or "newscontent" in lower
            or "content.jsp" in lower
            or "wbnewsid=" in lower
            or "wbnewsid" in href.lower()
        )
        if not is_content:
            continue
        if full_url in seen_urls:
            continue

        container = pick_row_container(anchor)
        row_text = normalize_space(container.get_text(" ", strip=True))
        publish_date = first_date(row_text)
        department = infer_department(container, title, publish_date)
        results.append(
            NoticeSummary(
                title=title,
                url=full_url,
                department=department,
                publish_date=publish_date,
            )
        )
        seen_urls.add(full_url)

    return results


def next_page_from_html(html: str, base_url: str) -> Optional[str]:
    soup = BeautifulSoup(html, "html.parser")
    wanted_text = {"下页", "下一页", ">", "›"}
    for anchor in soup.find_all("a", href=True):
        text = normalize_space(anchor.get_text(" ", strip=True))
        title = normalize_space(anchor.get("title", ""))
        if text not in wanted_text and "下一页" not in title and "下页" not in title:
            continue
        href = anchor.get("href", "").strip()
        if not href or href == "#" or href.lower().startswith("javascript:"):
            continue
        return urljoin(base_url, href)
    return None


def click_next_page(page) -> bool:
    html = page.content()
    next_url = next_page_from_html(html, page.url)
    if next_url:
        safe_goto(page, next_url)
        return True

    selectors = [
        "a:has-text('下一页')",
        "a:has-text('下页')",
        "a:has-text('>')",
        "text=下一页",
        "text=下页",
    ]
    for selector in selectors:
        try:
            locator = page.locator(selector).last
            if locator.count() == 0:
                continue
            locator.click(timeout=3000)
            try:
                page.wait_for_load_state("domcontentloaded", timeout=10000)
                page.wait_for_load_state("networkidle", timeout=10000)
            except PlaywrightTimeoutError:
                pass
            return True
        except Exception:
            continue
    return False


def extract_article_data(html: str, fallback_title: str = "") -> ArticleData:
    soup = BeautifulSoup(html, "html.parser")
    for tag in soup(["script", "style", "noscript", "iframe"]):
        tag.decompose()

    title = fallback_title
    for selector in [
        "h1",
        "h2",
        ".titlestyle",
        ".title",
        ".article-title",
        ".news-title",
        ".content-title",
    ]:
        node = soup.select_one(selector)
        if node:
            candidate = normalize_space(node.get_text(" ", strip=True))
            if len(candidate) >= 4:
                title = candidate
                break

    full_text = soup.get_text("\n", strip=True)
    publish_department = ""
    publish_date = ""
    dept_match = re.search(r"发布部门\s*[:：]\s*([^\n\r]+?)(?=\s+发布|\s+浏览|\n|$)", full_text)
    if dept_match:
        publish_department = clean_line(dept_match.group(1))
    date_match = re.search(
        r"发布时间\s*[:：]\s*(20\d{2}[-/.年]\d{1,2}[-/.月]\d{1,2}日?)",
        full_text,
    )
    if date_match:
        publish_date = normalize_date_text(date_match.group(1))

    body_node = None
    for selector in [
        "#vsb_content",
        "#vsb_content_2",
        ".v_news_content",
        ".article-content",
        ".news-content",
        ".content",
        ".main-content",
        ".wp_articlecontent",
    ]:
        body_node = soup.select_one(selector)
        if body_node and len(body_node.get_text(strip=True)) > 80:
            break
    if not body_node:
        body_node = soup.body or soup

    lines = get_text_lines(body_node)
    body_text = "\n".join(lines)
    return ArticleData(
        title=title,
        body_text=body_text,
        lines=lines,
        publish_department=publish_department,
        publish_date=publish_date,
    )


def trim_value(value: str, max_len: int = 120) -> str:
    value = clean_line(value)
    value = re.sub(r"^[：:\s]+", "", value)
    value = re.sub(r"\s*[。；;]$", "", value)
    value = re.sub(r"^(一|二|三|四|五|六|七|八|九|十|\d+)[、.．]\s*", "", value)
    if len(value) > max_len:
        pieces = re.split(r"[。；;]\s*", value)
        value = pieces[0] if pieces and pieces[0] else value[:max_len]
    return value[:max_len].strip()


def find_labeled_value(lines: Sequence[str], labels: Sequence[str]) -> str:
    label_pattern = "|".join(re.escape(label) for label in labels)
    for index, line in enumerate(lines):
        if any(skip in line for skip in ("发布时间", "浏览", "公示期")):
            continue
        match = re.search(rf"(?:{label_pattern})\s*[:：]\s*(.+)$", line)
        if match:
            value = trim_value(match.group(1))
            if value and not value.endswith("如下"):
                return value
        match = re.search(rf"(?:{label_pattern})\s*$", line)
        if match and index + 1 < len(lines):
            next_line = lines[index + 1]
            next_match = re.search(rf"(?:{label_pattern})\s*[:：]\s*(.+)$", next_line)
            value = trim_value(next_match.group(1) if next_match else next_line)
            if value:
                return value
    return ""


def extract_event_time(article: ArticleData) -> str:
    labels = (
        "活动时间",
        "比赛时间",
        "讲座时间",
        "培训时间",
        "会议时间",
        "宣讲时间",
        "举办时间",
        "报名时间",
        "时间安排",
        "日程安排",
    )
    value = find_labeled_value(article.lines, labels)
    if value:
        return value

    for line in article.lines:
        if "发布时间" in line or "公示期" in line:
            continue
        if "时间" in line and any(keyword in line for keyword in EVENT_KEYWORDS):
            match = re.search(r"时间\s*[:：]\s*(.+)$", line)
            if match:
                return trim_value(match.group(1))

    date_patterns = [
        r"20\d{2}年\d{1,2}月\d{1,2}日(?:[^\n，,。；;]{0,25})?",
        r"\d{1,2}月\d{1,2}日(?:[^\n，,。；;]{0,25})?",
    ]
    for pattern in date_patterns:
        for match in re.finditer(pattern, article.body_text):
            around = article.body_text[max(0, match.start() - 30) : match.end() + 30]
            if any(keyword in around for keyword in EVENT_KEYWORDS):
                return trim_value(match.group(0), max_len=80)
    return ""


def extract_location(article: ArticleData) -> str:
    labels = (
        "活动地点",
        "比赛地点",
        "讲座地点",
        "培训地点",
        "会议地点",
        "宣讲地点",
        "举办地点",
        "活动场地",
        "比赛场地",
        "地点",
        "场地",
    )
    value = find_labeled_value(article.lines, labels)
    if value:
        return value

    place_words = ("校区", "教室", "会议室", "报告厅", "体育馆", "图书馆", "大厅", "礼堂", "楼", "馆")
    for line in article.lines:
        if len(line) > 100:
            continue
        if any(word in line for word in place_words) and any(
            keyword in line for keyword in EVENT_KEYWORDS
        ):
            return trim_value(line)
    return ""


def extract_department(article: ArticleData, fallback: str = "") -> str:
    labels = (
        "主办单位",
        "承办单位",
        "协办单位",
        "组织单位",
        "举办单位",
        "牵头单位",
        "负责部门",
    )
    value = find_labeled_value(article.lines, labels)
    if value:
        return value
    return article.publish_department or fallback


def parse_sort_time(raw_time: str, publish_date: str = "") -> str:
    year_hint = None
    if publish_date:
        match = re.match(r"(20\d{2})", publish_date)
        if match:
            year_hint = int(match.group(1))
    if year_hint is None:
        year_hint = dt.datetime.now().year

    text = raw_time.replace("：", ":")
    patterns = [
        r"(?P<year>20\d{2})年(?P<month>\d{1,2})月(?P<day>\d{1,2})日?"
        r"(?P<tail>[^\n。；;]{0,30})",
        r"(?P<year>20\d{2})[-/.](?P<month>\d{1,2})[-/.](?P<day>\d{1,2})"
        r"(?P<tail>[^\n。；;]{0,30})",
        r"(?P<month>\d{1,2})月(?P<day>\d{1,2})日(?P<tail>[^\n。；;]{0,30})",
    ]

    for pattern in patterns:
        match = re.search(pattern, text)
        if not match:
            continue
        year = int(match.groupdict().get("year") or year_hint)
        month = int(match.group("month"))
        day = int(match.group("day"))
        tail = match.groupdict().get("tail") or ""
        hour = 0
        minute = 0
        time_match = re.search(r"(?P<hour>\d{1,2})\s*(?::|点)\s*(?P<minute>\d{1,2})?", tail)
        if time_match:
            hour = int(time_match.group("hour"))
            minute = int(time_match.group("minute") or 0)
            if any(word in tail for word in ("下午", "晚上", "晚间")) and hour < 12:
                hour += 12
        try:
            return dt.datetime(year, month, day, hour, minute).strftime("%Y-%m-%d %H:%M")
        except ValueError:
            continue
    return ""


def is_on_campus_location(location: str) -> bool:
    text = normalize_space(location).lower()
    if not text:
        return False

    strong_campus_words = ("北京邮电大学", "北邮", "bupt")
    building_pattern = (
        r"(教[一二三四五六七八九十\d]+|教学楼|主楼|科研楼|经管楼|行政楼|"
        r"学生发展中心|学生活动中心|科学会堂|小西天|西土城|沙河|宏福|校区|校内)"
    )
    has_strong_campus = any(word.lower() in text for word in strong_campus_words)
    has_physical_campus = bool(re.search(building_pattern, location))
    has_non_campus = any(word.lower() in text for word in NON_CAMPUS_LOCATION_KEYWORDS)

    if has_non_campus and not has_physical_campus:
        return False
    if has_strong_campus or has_physical_campus:
        return True
    return False


def is_activity(article: ArticleData, notice: NoticeSummary, event_time: str, location: str, include_all: bool) -> bool:
    if include_all:
        return True
    text = f"{notice.title}\n{article.title}\n{article.body_text[:1000]}"
    keyword_hit = any(keyword in text for keyword in EVENT_KEYWORDS)
    if location and (event_time or keyword_hit):
        return True
    if event_time and keyword_hit:
        return True
    return False


def parse_activity(
    notice: NoticeSummary,
    html: str,
    include_all: bool,
    campus_only: bool = True,
    target_date: str = "",
) -> Optional[ActivityRow]:
    article = extract_article_data(html, notice.title)
    event_time = extract_event_time(article)
    location = extract_location(article)
    department = extract_department(article, notice.department)
    publish_date = article.publish_date or notice.publish_date

    if target_date and publish_date != target_date:
        return None
    if not is_activity(article, notice, event_time, location, include_all):
        return None
    if campus_only and not is_on_campus_location(location):
        return None

    note_parts = []
    if not location:
        note_parts.append("未识别到地点")
    if not event_time:
        note_parts.append("未识别到活动时间")
    if not department:
        note_parts.append("未识别到部门")

    sort_time = parse_sort_time(event_time, publish_date)
    return ActivityRow(
        location=location,
        event_time=event_time,
        department=department,
        activity_name=article.title or notice.title,
        publish_date=publish_date,
        notice_title=notice.title,
        url=notice.url,
        note="；".join(note_parts),
        sort_time=sort_time,
    )


def collect_notices(page, list_url: str, pages: int, max_items: int) -> List[NoticeSummary]:
    safe_goto(page, list_url)
    all_notices: List[NoticeSummary] = []
    seen = set()

    for page_index in range(1, pages + 1):
        html = page.content()
        notices = extract_notice_summaries(html, page.url)
        added = 0
        for notice in notices:
            if notice.url in seen:
                continue
            seen.add(notice.url)
            all_notices.append(notice)
            added += 1
            if len(all_notices) >= max_items:
                break
        log(f"列表第 {page_index} 页提取 {added} 条新通知，累计 {len(all_notices)} 条。")
        if len(all_notices) >= max_items or page_index >= pages:
            break
        if not click_next_page(page):
            log("没有找到下一页按钮，停止翻页。")
            break
    return all_notices[:max_items]


def crawl(args) -> List[ActivityRow]:
    chrome_path = find_chrome_executable(args.chrome_path)
    if chrome_path:
        log(f"使用浏览器：{chrome_path}")
    else:
        log("没有找到系统 Chrome/Edge，将尝试使用 Playwright 自带浏览器。")

    profile_dir = Path(args.profile_dir).expanduser().resolve()
    profile_dir.mkdir(parents=True, exist_ok=True)

    with sync_playwright() as playwright:
        context = playwright.chromium.launch_persistent_context(
            user_data_dir=str(profile_dir),
            executable_path=chrome_path,
            headless=args.headless,
            locale="zh-CN",
            viewport={"width": 1440, "height": 1000},
            slow_mo=args.slow_mo,
            args=[
                "--disable-blink-features=AutomationControlled",
                "--disable-dev-shm-usage",
            ],
        )
        page = context.pages[0] if context.pages else context.new_page()
        page.set_default_timeout(args.timeout * 1000)

        page = ensure_logged_in(page, args.list_url, args.login_timeout)
        notices = collect_notices(page, args.list_url, args.pages, args.max_items)
        if args.target_date:
            log(f"仅整理发布日期为 {args.target_date} 的通知。")

        rows: List[ActivityRow] = []
        for index, notice in enumerate(notices, start=1):
            if args.target_date and notice.publish_date and notice.publish_date != args.target_date:
                continue
            log(f"读取正文 {index}/{len(notices)}：{notice.title}")
            try:
                safe_goto(page, notice.url, timeout_ms=args.timeout * 1000)
                if is_login_page(page):
                    page = ensure_logged_in(page, notice.url, args.login_timeout)
                row = parse_activity(
                    notice,
                    page.content(),
                    args.include_all,
                    campus_only=not args.include_off_campus,
                    target_date=args.target_date,
                )
                if row:
                    rows.append(row)
            except Exception as exc:
                if args.target_date and notice.publish_date != args.target_date:
                    continue
                rows.append(
                    ActivityRow(
                        location="",
                        event_time="",
                        department=notice.department,
                        activity_name=notice.title,
                        publish_date=notice.publish_date,
                        notice_title=notice.title,
                        url=notice.url,
                        note=f"抓取失败：{exc}",
                        sort_time="",
                    )
                )
            if args.delay > 0:
                time.sleep(args.delay)

        context.close()
    rows.sort(
        key=lambda row: (
            row.location or "~~~~",
            row.sort_time or "9999-12-31 23:59",
            row.department or "~~~~",
            row.activity_name,
        )
    )
    return rows


def write_excel(rows: Sequence[ActivityRow], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "活动整理"

    headers = [
        "活动地点",
        "活动时间",
        "部门",
        "活动名称",
        "发布时间",
        "通知标题",
        "原文链接",
        "备注",
    ]
    ws.append(headers)
    for row in rows:
        ws.append(
            [
                row.location,
                row.event_time,
                row.department,
                row.activity_name,
                row.publish_date,
                row.notice_title,
                row.url,
                row.note,
            ]
        )

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    widths = [34, 28, 28, 46, 14, 46, 70, 24]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    if rows:
        table_ref = f"A1:H{ws.max_row}"
        table = Table(displayName="ActivityTable", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)
    else:
        ws.append(["未抽取到符合条件的校内活动。可增加 --pages，或使用 --include-off-campus 放宽地点筛选。"])
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))

    meta = wb.create_sheet("运行说明")
    meta.append(["字段", "说明"])
    meta.append(["排序", "已按 活动地点 -> 活动时间 -> 部门 排序"])
    meta.append(["登录态", "首次登录后保存在 runtime/chrome-profile，不保存账号密码"])
    meta.append(["备注", "未识别字段会在备注列提示，建议人工复核关键活动"])
    for cell in meta[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    meta.column_dimensions["A"].width = 18
    meta.column_dimensions["B"].width = 80
    meta.freeze_panes = "A2"

    wb.save(output_path)


def positive_int(value: str) -> int:
    parsed = int(value)
    if parsed <= 0:
        raise argparse.ArgumentTypeError("必须是正整数")
    return parsed


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="自动登录北邮信息门户，抓取校内通知并整理活动 Excel。"
    )
    parser.add_argument("--list-url", default=DEFAULT_LIST_URL, help="校内通知列表页 URL")
    parser.add_argument("--home-url", default=DEFAULT_HOME_URL, help="信息门户首页 URL")
    parser.add_argument("--pages", type=positive_int, default=3, help="抓取列表页数，默认 3")
    parser.add_argument(
        "--max-items", type=positive_int, default=120, help="最多读取通知数量，默认 120"
    )
    parser.add_argument(
        "--output",
        default=str(default_output_path()),
        help="Excel 输出路径，默认 outputs/校内通知活动_时间戳.xlsx",
    )
    parser.add_argument(
        "--profile-dir",
        default=str(app_dir() / "runtime" / "chrome-profile"),
        help="浏览器登录态保存目录",
    )
    parser.add_argument("--chrome-path", default="", help="Chrome/Edge 可执行文件路径")
    parser.add_argument(
        "--login-timeout", type=positive_int, default=180, help="等待手动登录秒数"
    )
    parser.add_argument("--timeout", type=positive_int, default=60, help="页面加载超时秒数")
    parser.add_argument("--delay", type=float, default=0.5, help="读取正文间隔秒数")
    parser.add_argument("--slow-mo", type=int, default=0, help="浏览器操作慢放毫秒数")
    parser.add_argument("--headless", action="store_true", help="无头模式，仅适合已有登录态")
    parser.add_argument(
        "--today-only",
        action="store_true",
        help="只整理运行当天发布的通知",
    )
    parser.add_argument(
        "--date",
        default="",
        help="只整理指定发布日期的通知，格式 YYYY-MM-DD",
    )
    parser.add_argument(
        "--include-all",
        action="store_true",
        help="输出全部通知，而不是只输出疑似活动通知",
    )
    parser.add_argument(
        "--include-off-campus",
        action="store_true",
        help="放宽地点筛选，允许输出线上、校外或地点不明确的通知",
    )
    return parser


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)

    try:
        output_path = Path(args.output).expanduser().resolve()
        args.target_date = resolve_target_date(args)
        rows = crawl(args)
        write_excel(rows, output_path)
        log(f"完成：共输出 {len(rows)} 条活动。")
        log(f"Excel 文件：{output_path}")
        return 0
    except KeyboardInterrupt:
        print("\n已取消。", file=sys.stderr)
        return 130
    except Exception as exc:
        print(f"运行失败：{exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
